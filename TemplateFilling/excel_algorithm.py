import json
import openpyxl as oxl

template = oxl.load_workbook("templates/excel/Templates/1_template_standardExcel/input_standardExcel-Template.xlsx")

data = json.loads(open("templates/excel/Templates/1_template_standardExcel/input_standardExcel-Data.json", "r").read())

print(template.sheetnames)
print(data)


wb = oxl.Workbook()
ws0 = wb.active
ws0.title = template.sheetnames[0]

i=1
for sheet in template.sheetnames[1:]:
    
       
    # Sheets can be added to workbook with the
    # workbook object's create_sheet() method. 
    wb.create_sheet(index = i , title = sheet)

    i+=1
  
wb.save("try.xlsx")