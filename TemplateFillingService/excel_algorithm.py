import json
import openpyxl as oxl

# load template
template = oxl.load_workbook("templates/excel/Templates/1_template_standardExcel/input_standardExcel-Template.xlsx")

# load json data
data = json.loads(open("templates/excel/Templates/1_template_standardExcel/input_standardExcel-Data.json", "r").read())

i = 0
# go through all sheets
for sheet in template.sheetnames:

    if i > 0 and sheet not in template:
        # create new sheet
        ws = template.create_sheet(index=i, title=sheet)
    else:
        ws = template[sheet]

    # fill sheet with data
    for cell in data[i][sheet]:
        cell_number = list(cell.keys())
        ws[cell_number[0]] = cell[cell_number[0]]

    i += 1

template.save("try.xlsx")
